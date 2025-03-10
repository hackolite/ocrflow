import requests
import xlsxwriter
from openpyxl import load_workbook
from datetime import datetime



service_url = "prices.openfoodfacts.org"



# Obtenir la date actuelle
def get_now():
	now = datetime.now()
	# Formater la date selon le format "YYYY-MM-DD"
	date_formatted = now.strftime("%Y-%m-%d")
	# Date arbitraire
	#date_arbitraire = "2024-03-21"

	# Convertir la date arbitraire en objet datetime
	date_object = datetime.strptime(date_formatted, "%Y-%m-%d")
	date_object = date_object.strftime("%Y-%m-%d")
	return date_object


def ingest(excel=None):
	token = excel
	return token




def send_image_proof(data_type='PRICE_TAG', headers=None, image_path=None):
	url = 'https://{}/api/v1/proofs/upload'.format(service_url)
	print(url)
	files = {'file': (image_path, open(image_path, 'rb'), 'image/jpeg')}
	data = {'type': data_type}
	response = requests.post(url, headers=headers, data=data, files=files)
	print("proof status:", response.reason)
	return response.json()



def send_product(headers, ean, price, proof, osm_id=None):
	date_formatted = get_now()
	url = "https://{}/api/v1/prices".format(service_url)
	data = {"date":date_formatted, "product_code": str(int(float(ean))),"price": price,"currency": "EUR","location_osm_id": 145131721 ,"location_osm_type": "WAY","proof_id": proof}
	print(data)
	response = requests.post(url, headers=headers, json=data)
	return response


def connection(user, password):

	headers = {'accept': 'application/json','Content-Type': 'application/x-www-form-urlencoded'}

	data = {
		'set_cookie':'true',
		    'grant_type': '',
	    'username': user,
	    'password': password,
	    'scope': '',
	    'client_id': '',
	    'client_secret': ''
	}

	url = "https://{}/api/v1/auth".format(service_url)
	token = requests.post(url, data=data, headers=headers)
	print("token", token)
	return token

def send(folder):
	resp = connection("laureote", "")

	token = resp.json()['access_token']
	headers = {'Authorization': 'Bearer {}'.format(token)}

	workbook = load_workbook(filename="{}/tmp.xlsx".format(folder))
	sheet = workbook['Sheet1']

	for row in sheet.iter_rows(values_only=True):
			try:
				ean = str(row[1])
				price = str(row[2])
				path = "{}/{}".format(folder,str(row[7]))
				if "=TRUE()" in  row[4]:
					print("SENDING")
					print(path)
					resp = send_image_proof(headers=headers, image_path=path)
					print("resp")
					proof = resp['id']
					resp = send_product(headers, ean, price, proof, osm_id=10711242347)
					print(resp.json())
					print("send product status code:", resp.status_code)
			except Exception as e:
				print(e)


def delete_product():
	resp = connection("laureote", "")
	token = resp.json()['access_token']
	headers = {'Authorization': 'Bearer {}'.format(token)}
    
	values = range(4518,4596)
	for price_id in values:
			price_id = str(price_id)
			url = "https://{}/api/v1/prices/{}".format(service_url, price_id)
			response = requests.delete(url, headers=headers)
			print(response.status_code)

send("../cora")
